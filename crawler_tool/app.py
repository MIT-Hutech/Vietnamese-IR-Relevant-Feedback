import io
import json
import logging
import os
import queue
import threading
import time
from datetime import datetime
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request, send_file
from flask_cors import CORS

import crawler as crawl_module
from crawler import DOMAIN_CONFIG

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                                  numbers)
except ImportError:
    openpyxl = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# ── In-memory state ────────────────────────────────────────────────────────────
REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = REPOSITORY_ROOT / "data" / "dataset.json"
DATA_FILE.parent.mkdir(exist_ok=True)

_state = {
    "running": False,
    "records": [],        # accumulated records
    "progress": 0,
    "target": 1500,
    "log_queue": queue.Queue(maxsize=500),
    "thread": None,
    "started_at": None,
    "finished": False,
    "error": None,
}


def _load_saved():
    if DATA_FILE.exists():
        try:
            with open(DATA_FILE, encoding="utf-8") as f:
                _state["records"] = json.load(f)
                _state["progress"] = len(_state["records"])
        except Exception:
            pass


def _save_data():
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(_state["records"], f, ensure_ascii=False, indent=2)


def _crawl_task(target: int, resume: bool):
    _state["running"] = True
    _state["finished"] = False
    _state["error"] = None
    _state["started_at"] = datetime.now().isoformat()

    try:
        def progress_cb(current, total, msg):
            _state["progress"] = len(_state["records"])
            try:
                _state["log_queue"].put_nowait(
                    {"type": "progress", "current": len(_state["records"]), "total": total, "msg": msg}
                )
            except queue.Full:
                pass

        # Pass existing titles to skip on resume
        existing_titles = {r.get("title", "") for r in _state["records"]}

        for record in crawl_module.crawl_all(target=target, progress_cb=progress_cb,
                                              existing_titles=existing_titles):
            _state["records"].append(record)
            _state["progress"] = len(_state["records"])
            # Auto-save every 50 records
            if _state["progress"] % 50 == 0:
                _save_data()
                logger.info(f"Saved checkpoint: {_state['progress']} records")

        _save_data()
        _state["finished"] = True
        _state["log_queue"].put_nowait(
            {"type": "done", "current": _state["progress"], "total": target}
        )
    except Exception as e:
        _state["error"] = str(e)
        logger.exception("Crawl error")
        _state["log_queue"].put_nowait({"type": "error", "msg": str(e)})
    finally:
        _state["running"] = False


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", domains=DOMAIN_CONFIG)


@app.route("/api/status")
def api_status():
    domain_counts = {}
    for r in _state["records"]:
        d = r.get("domain", "Unknown")
        domain_counts[d] = domain_counts.get(d, 0) + 1

    return jsonify({
        "running": _state["running"],
        "finished": _state["finished"],
        "progress": _state["progress"],
        "target": _state["target"],
        "total_records": len(_state["records"]),
        "domain_counts": domain_counts,
        "started_at": _state["started_at"],
        "error": _state["error"],
    })


@app.route("/api/start", methods=["POST"])
def api_start():
    if _state["running"]:
        return jsonify({"error": "Already running"}), 400

    body = request.get_json(silent=True) or {}
    target = int(body.get("target", 1500))
    resume = bool(body.get("resume", True))

    if not resume:
        _state["records"] = []
        _state["progress"] = 0

    _state["target"] = target
    # Clear log queue
    while not _state["log_queue"].empty():
        try:
            _state["log_queue"].get_nowait()
        except queue.Empty:
            break

    t = threading.Thread(target=_crawl_task, args=(target, resume), daemon=True)
    _state["thread"] = t
    t.start()

    return jsonify({"ok": True, "target": target})


@app.route("/api/stop", methods=["POST"])
def api_stop():
    # Signal by setting running flag — crawler checks via generator exhaustion
    _state["running"] = False
    _save_data()
    return jsonify({"ok": True, "saved": len(_state["records"])})


@app.route("/api/stream")
def api_stream():
    """SSE stream for real-time progress updates."""
    def event_gen():
        while True:
            try:
                msg = _state["log_queue"].get(timeout=2)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("type") in ("done", "error"):
                    break
            except queue.Empty:
                # Heartbeat
                yield f"data: {json.dumps({'type': 'heartbeat', 'progress': _state['progress']})}\n\n"
                if not _state["running"] and _state["finished"]:
                    break

    return Response(event_gen(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/preview")
def api_preview():
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))
    domain_filter = request.args.get("domain", "")
    query_filter = request.args.get("q", "").lower()

    records = _state["records"]
    if domain_filter:
        records = [r for r in records if r.get("domain") == domain_filter]
    if query_filter:
        records = [r for r in records if query_filter in r.get("title", "").lower()
                   or query_filter in r.get("query", "").lower()]

    total = len(records)
    start = (page - 1) * per_page
    end = start + per_page
    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "records": records[start:end]
    })


@app.route("/api/export/xlsx")
def api_export_xlsx():
    if not openpyxl:
        return jsonify({"error": "openpyxl not installed"}), 500

    records = _state["records"]
    if not records:
        return jsonify({"error": "No data to export"}), 400

    wb = openpyxl.Workbook()
    # ── Sheet 1: Main dataset ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dataset"

    headers = ["ID", "Lĩnh vực", "Tiêu đề", "Câu truy vấn",
               "Tài liệu (snippet)", "Số từ", "Độ liên quan (1-3)",
               "URL nguồn"]

    header_fill = PatternFill("solid", fgColor="1E1B4B")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    domain_colors = {
        "Khoa học": "EEF2FF",
        "Công nghệ": "F5F3FF",
        "Sức khỏe & Y tế": "FDF2F8",
        "Lịch sử": "FFFBEB",
        "Game & Giải trí": "ECFDF5",
        "Môi trường": "F0FDFA",
        "Kinh tế & Xã hội": "FFF7ED",
    }

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for row_idx, rec in enumerate(records, 2):
        domain = rec.get("domain", "")
        row_fill = PatternFill("solid", fgColor=domain_colors.get(domain, "FFFFFF"))
        values = [
            rec.get("id"),
            domain,
            rec.get("title"),
            rec.get("query"),
            rec.get("document", "")[:500],
            rec.get("word_count"),
            rec.get("relevance"),
            rec.get("url"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (4, 5)))
            if col == 6:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col == 7:
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Column widths
    col_widths = [6, 18, 40, 35, 80, 10, 15, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Domain summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Thống kê lĩnh vực")
    ws2.append(["Lĩnh vực", "Số tài liệu", "Trung bình số từ", "TB độ liên quan"])
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    domain_stats = {}
    for r in records:
        d = r.get("domain", "Unknown")
        if d not in domain_stats:
            domain_stats[d] = {"count": 0, "words": 0, "rel": 0}
        domain_stats[d]["count"] += 1
        domain_stats[d]["words"] += r.get("word_count", 0)
        domain_stats[d]["rel"] += r.get("relevance", 0)

    for row_i, (dom, stats) in enumerate(domain_stats.items(), 2):
        c = stats["count"]
        ws2.cell(row=row_i, column=1, value=dom)
        ws2.cell(row=row_i, column=2, value=c)
        ws2.cell(row=row_i, column=3, value=round(stats["words"] / c, 1) if c else 0)
        ws2.cell(row=row_i, column=4, value=round(stats["rel"] / c, 2) if c else 0)
        fill = PatternFill("solid", fgColor=domain_colors.get(dom, "FFFFFF"))
        for col in range(1, 5):
            ws2.cell(row=row_i, column=col).fill = fill

    for col_letter, w in zip(["A", "B", "C", "D"], [25, 15, 20, 20]):
        ws2.column_dimensions[col_letter].width = w

    # ── Sheet 3: IR Relevance Feedback schema ─────────────────────────────────
    ws3 = wb.create_sheet("IR Relevant Feedback")
    ws3.append([
        "query_id", "query_text", "doc_id", "doc_title",
        "relevance_label", "feedback_type", "domain", "note"
    ])
    for col in range(1, 9):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    feedback_map = {1: "Non-relevant", 2: "Partially relevant", 3: "Highly relevant"}
    # Group by query to build qrel pairs
    query_groups = {}
    for rec in records:
        q = rec.get("query", "")
        if q not in query_groups:
            query_groups[q] = []
        query_groups[q].append(rec)

    row_i = 2
    for q_id, (query_text, docs) in enumerate(query_groups.items(), 1):
        for doc in docs:
            ws3.append([
                f"Q{q_id:04d}",
                query_text,
                f"D{doc['id']:05d}",
                doc.get("title"),
                doc.get("relevance"),
                feedback_map.get(doc.get("relevance", 1), "Non-relevant"),
                doc.get("domain"),
                ""  # empty note for manual annotation
            ])
            row_i += 1

    for col_letter, w in zip(["A","B","C","D","E","F","G","H"], [10,40,10,40,15,20,20,20]):
        ws3.column_dimensions[col_letter].width = w

    # ── Export ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ir_dataset_vi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/api/export/json")
def api_export_json():
    buf = io.BytesIO(json.dumps(_state["records"], ensure_ascii=False, indent=2).encode("utf-8"))
    buf.seek(0)
    filename = f"ir_dataset_vi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(buf, mimetype="application/json", as_attachment=True, download_name=filename)


if __name__ == "__main__":
    _load_saved()
    logger.info(f"Loaded {len(_state['records'])} existing records")
    app.run(debug=False, port=5050, threaded=True)
